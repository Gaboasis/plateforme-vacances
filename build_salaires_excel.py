"""Génère salaires_2025-2027.xlsx : % d'augm. modifiable → recalcule 2026-2027.

Formules en anglais (ROUND, SUM) pour Excel macOS US / interface anglaise.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/stafa/Plateforme Vacances/salaires_2025-2027.xlsx"

# Dossier, Nom, Catégorie, Taux 25-26, Hres/année, Taux 26-27 (pour % initial seulement)
ROWS = [
    (1507, "Al Chafei Amneh", "Cuisinière", 18.36, 1440, 19.28),
    (657, "Bachtaka, Saliha", "Éducatrice qualifiée", 25.50, 1600, 26.78),
    (81, "Bouazza Azza", "Éducatrice non qualifiée", 19.00, 1750, 19.95),
    (826, "Cheikh-El-Najjarine, Hanady", "Éducatrice non qualifiée", 21.00, 1840, 22.05),
    (203, "Cherbib, Rafika", "Éducatrice qualifiée", 25.00, 1560, 26.25),
    (1251, "ElAlti Loubaba", "Éducatrice non qualifiée", 21.00, 1750, 22.05),
    (1657, "El Bhetouri Karima", "Éducatrice qualifiée", 25.00, 1750, 26.25),
    (441, "Ghazzawi, Rana", "Éducatrice qualifiée", 27.60, 1750, 28.98),
    (1666, "Haidar Ahmad Manal", "Educatrice non qualifiée", 19.00, 1540, 19.95),
    (1576, "Hemiss Om El Khir", "Éducatrice non qualifiée", 20.00, 800, 21.00),
    (560, "Kendakji, Zoukaa", "Préposée", 18.00, 1760, 18.90),
    (1550, "Noussir Hajar", "Remplaçante", 18.25, 1125, 19.16),
    (759, "Ramadan, Kamar", "Adjointe administrative", 21.50, 1680, 22.58),
    (1370, "Romdhane, Souhir", "Éducatrice qualifiée", 26.00, 1400, 27.30),
]

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill_2526 = PatternFill("solid", fgColor="E2EFDA")
header_fill_2627 = PatternFill("solid", fgColor="BDD7EE")
header_fill_misc = PatternFill("solid", fgColor="F2F2F2")
header_fill_pct = PatternFill("solid", fgColor="FFFACD")
total_fill = PatternFill("solid", fgColor="FFF2CC")

NCOLS = 10


def pct_initial(t2526, t2627):
    return round((t2627 / t2526 - 1) * 100, 2)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Salaires"

    ws.merge_cells("D1:F1")
    ws["D1"] = "2025-2026"
    ws["G1"] = "Augmentation %\n(modifiable)"
    ws.merge_cells("H1:I1")
    ws["H1"] = "2026-2027"
    ws["J1"] = "Augmentation $"

    ws["D1"].font = Font(bold=True)
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D1"].fill = header_fill_2526

    ws["G1"].font = Font(bold=True)
    ws["G1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G1"].fill = header_fill_pct

    ws["H1"].font = Font(bold=True)
    ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H1"].fill = header_fill_2627

    ws["J1"].font = Font(bold=True)
    ws["J1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["J1"].fill = header_fill_misc

    h2 = (
        "Dossier",
        "Nom",
        "Catégorie",
        "Taux horaire",
        "Hres/année",
        "Salaire/année",
        "Hausse\n(%)",
        "Taux horaire",
        "Salaire/année",
        "26-27 vs 25-26",
    )
    for col, h in enumerate(h2, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col <= 3:
            c.fill = header_fill_misc
        elif col <= 6:
            c.fill = header_fill_2526
        elif col == 7:
            c.fill = header_fill_pct
        elif col <= 9:
            c.fill = header_fill_2627
        else:
            c.fill = header_fill_misc

    first_data = 3
    last_data = first_data + len(ROWS) - 1

    for r, row in enumerate(ROWS, start=first_data):
        dossier, nom, cat, t2526, hres, t2627 = row
        ws.cell(row=r, column=1, value=dossier)
        ws.cell(row=r, column=2, value=nom)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=t2526).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5, value=hres).number_format = "0"
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7, value=pct_initial(t2526, t2627)).number_format = "0.00"
        # Taux 26-27 = taux 25-26 × (1 + %/100), arrondi 2 déc. (ROUND = Excel EN)
        ws.cell(row=r, column=8, value=f"=ROUND(D{r}*(1+G{r}/100),2)").number_format = (
            '"$"#,##0.00'
        )
        ws.cell(row=r, column=9, value=f"=ROUND(H{r}*E{r},2)").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=10, value=f"=I{r}-F{r}").number_format = '"$"#,##0.00'

    tr = last_data + 1
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=1).font = Font(bold=True)
    ws.cell(row=tr, column=1).fill = total_fill
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="left")

    for col in range(2, NCOLS + 1):
        cell = ws.cell(row=tr, column=col)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        if col == 6:
            cell.value = f"=SUM(F{first_data}:F{last_data})"
            cell.number_format = '"$"#,##0.00'
        elif col == 9:
            cell.value = f"=SUM(I{first_data}:I{last_data})"
            cell.number_format = '"$"#,##0.00'
        elif col == 10:
            cell.value = f"=SUM(J{first_data}:J{last_data})"
            cell.number_format = '"$"#,##0.00'

    for row in ws.iter_rows(min_row=1, max_row=tr, min_col=1, max_col=NCOLS):
        for cell in row:
            cell.border = border

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24
    for col in range(4, NCOLS + 1):
        if col == 7:
            w = 11
        elif col in (8, 9, 10):
            w = 15
        else:
            w = 14
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 40
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
